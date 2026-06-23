"""项目档案 Excel 导入导出，支持单项目和整库迁移。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from database import Database
from services.file_service import FileService


PROJECT_HEADERS = [
    ("id", "项目ID"),
    ("parent_project_id", "上级项目ID"),
    ("name", "项目名称"),
    ("short_name", "子项目简称"),
    ("construction_company", "施工单位"),
    ("contract_amount", "合同金额"),
    ("design_company", "设计单位"),
    ("design_amount", "设计金额"),
    ("supervision_company", "监理单位"),
    ("supervision_amount", "监理金额"),
    ("start_date", "开工日期"),
    ("completion_date", "完工日期"),
    ("contact_person", "联系人"),
    ("contact_phone", "联系电话"),
    ("remarks", "备注"),
    ("created_at", "创建时间"),
]

FILE_HEADERS = [
    ("file_id", "档案ID"),
    ("project_id", "项目ID"),
    ("project_name", "项目名称"),
    ("box_no", "盒号"),
    ("file_name", "文件名称"),
    ("file_type", "载体类型"),
    ("count_text", "份数描述"),
    ("original_count", "原件库存"),
    ("copy_count", "复印件库存"),
    ("file_path", "电子文件路径"),
    ("deleted_at", "删除时间"),
    ("computed_status", "当前状态"),
    ("borrowers", "当前借用人"),
]

BORROW_HEADERS = [
    ("id", "借阅ID"),
    ("file_id", "档案ID"),
    ("borrower", "借用人"),
    ("contact", "联系电话"),
    ("reason", "事由"),
    ("media_type", "借阅类型"),
    ("quantity", "借阅份数"),
    ("borrow_date", "借出日期"),
    ("expected_return_date", "预计归还日期"),
    ("return_date", "实际归还日期"),
]

IMPORT_ALIASES = {
    "盒号": "box_no",
    "文件名称": "name",
    "载体类型": "type",
    "份数描述": "count_text",
    "原件库存": "original_count",
    "复印件库存": "copy_count",
    "电子文件路径": "file_path",
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


class ExcelService:
    def __init__(self, database: Database, file_service: FileService) -> None:
        self.database = database
        self.file_service = file_service

    @staticmethod
    def _style_sheet(sheet: Worksheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            values = [len(_text(cell.value)) for cell in column]
            width = min(max(max(values, default=8) + 2, 10), 42)
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _write_sheet(
        workbook: Workbook,
        title: str,
        headers: list[tuple[str, str]],
        rows: Iterable[dict[str, Any]],
    ) -> Worksheet:
        sheet = workbook.create_sheet(title)
        sheet.append([label for _, label in headers])
        for row in rows:
            sheet.append([row.get(field) for field, _ in headers])
        ExcelService._style_sheet(sheet)
        return sheet

    def export_project(self, project_id: int, output_path: str | Path) -> Path:
        project = self.file_service.get_project(project_id)
        if project is None:
            raise LookupError("项目不存在或已被删除。")
        output = Path(output_path).expanduser().resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        rows = self.file_service.list_files(project_id)
        self._write_sheet(workbook, "档案台账", FILE_HEADERS, rows)
        self._write_sheet(workbook, "项目资料", PROJECT_HEADERS, [project])
        file_ids = [row["file_id"] for row in rows]
        borrows = self._borrows_for_files(file_ids)
        self._write_sheet(workbook, "借阅历史", BORROW_HEADERS, borrows)
        workbook.save(output)
        return output

    def export_all(self, output_path: str | Path) -> Path:
        output = Path(output_path).expanduser().resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        projects = self.file_service.list_projects()
        files = self.database.fetch_all(
            """
            SELECT
                f.id AS file_id, f.project_id, p.name AS project_name,
                f.box_no, f.name AS file_name, f.type AS file_type,
                f.count_text, f.original_count, f.copy_count, f.file_path,
                f.deleted_at,
                CASE WHEN f.deleted_at IS NULL THEN '正常' ELSE '回收站' END
                    AS computed_status,
                '' AS borrowers
            FROM Files AS f
            INNER JOIN Projects AS p ON p.id = f.project_id
            ORDER BY f.project_id, f.id
            """
        )
        borrows = self.database.fetch_all(
            "SELECT * FROM BorrowRecords ORDER BY id"
        )
        self._write_sheet(workbook, "项目", PROJECT_HEADERS, projects)
        self._write_sheet(workbook, "档案", FILE_HEADERS, files)
        self._write_sheet(workbook, "借阅历史", BORROW_HEADERS, borrows)
        guide = workbook.create_sheet("导入说明")
        guide.append(["项目档案本地管理系统 V0.2 整库迁移文件"])
        guide.append(["请勿修改各表中的 ID；导入会替换当前全部数据。"])
        guide.column_dimensions["A"].width = 72
        guide["A1"].font = Font(size=16, bold=True, color="1F4E78")
        workbook.save(output)
        return output

    def import_project(self, project_id: int, input_path: str | Path) -> dict[str, Any]:
        if self.file_service.get_project(project_id) is None:
            raise LookupError("项目不存在或已被删除。")
        workbook = load_workbook(
            Path(input_path).expanduser().resolve(), read_only=True, data_only=True
        )
        sheet = workbook[workbook.sheetnames[0]]
        headers = {_text(cell.value): index for index, cell in enumerate(sheet[1])}
        if "文件名称" not in headers:
            raise ValueError("Excel 至少需要包含“文件名称”列。")
        imported = 0
        errors: list[str] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            values: dict[str, Any] = {"project_id": project_id}
            for excel_name, field_name in IMPORT_ALIASES.items():
                index = headers.get(excel_name)
                values[field_name] = row[index] if index is not None else ""
            if not any(_text(value) for key, value in values.items() if key != "project_id"):
                continue
            values["original_count"] = values.get("original_count") or 1
            values["copy_count"] = values.get("copy_count") or 0
            try:
                self.file_service.create_file(values)
                imported += 1
            except Exception as exc:
                errors.append(f"第 {row_number} 行：{exc}")
        workbook.close()
        return {"imported": imported, "failed": len(errors), "errors": errors}

    def import_all(self, input_path: str | Path) -> dict[str, int]:
        input_file = Path(input_path).expanduser().resolve()
        if not input_file.is_file():
            raise FileNotFoundError("Excel 文件不存在。")
        workbook = load_workbook(input_file, read_only=True, data_only=True)
        required = {"项目", "档案", "借阅历史"}
        if required - set(workbook.sheetnames):
            raise ValueError("整库 Excel 必须包含“项目”“档案”“借阅历史”工作表。")
        project_rows = self._sheet_dicts(workbook["项目"])
        file_rows = self._sheet_dicts(workbook["档案"])
        borrow_rows = self._sheet_dicts(workbook["借阅历史"])
        workbook.close()

        project_by_label = {label: field for field, label in PROJECT_HEADERS}
        file_by_label = {label: field for field, label in FILE_HEADERS}
        borrow_by_label = {label: field for field, label in BORROW_HEADERS}
        with self.database.connection() as conn:
            conn.execute("DELETE FROM BorrowRecords")
            conn.execute("DELETE FROM Files")
            conn.execute("DELETE FROM Projects")
            for raw in project_rows:
                row = {project_by_label[key]: value for key, value in raw.items() if key in project_by_label}
                conn.execute(
                    """
                    INSERT INTO Projects(
                        id, parent_project_id, name, short_name,
                        construction_company, contract_amount,
                        design_company, design_amount,
                        supervision_company, supervision_amount,
                        start_date, completion_date, contact_person,
                        contact_phone, remarks, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    tuple(row.get(field) for field, _ in PROJECT_HEADERS),
                )
            for raw in file_rows:
                row = {file_by_label[key]: value for key, value in raw.items() if key in file_by_label}
                conn.execute(
                    """
                    INSERT INTO Files(
                        id, project_id, box_no, name, type, count_text,
                        original_count, copy_count, file_path, deleted_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row.get("file_id"),
                        row.get("project_id"),
                        row.get("box_no"),
                        row.get("file_name"),
                        row.get("file_type"),
                        row.get("count_text"),
                        row.get("original_count") or 1,
                        row.get("copy_count") or 0,
                        row.get("file_path"),
                        row.get("deleted_at"),
                    ),
                )
            for raw in borrow_rows:
                row = {borrow_by_label[key]: value for key, value in raw.items() if key in borrow_by_label}
                conn.execute(
                    """
                    INSERT INTO BorrowRecords(
                        id, file_id, borrower, contact, reason, media_type,
                        quantity, borrow_date, expected_return_date, return_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    tuple(row.get(field) for field, _ in BORROW_HEADERS),
                )
        return {
            "projects": len(project_rows),
            "files": len(file_rows),
            "borrows": len(borrow_rows),
        }

    def _borrows_for_files(self, file_ids: list[int]) -> list[dict[str, Any]]:
        if not file_ids:
            return []
        placeholders = ", ".join("?" for _ in file_ids)
        return self.database.fetch_all(
            f"SELECT * FROM BorrowRecords WHERE file_id IN ({placeholders}) ORDER BY id",
            file_ids,
        )

    @staticmethod
    def _sheet_dicts(sheet: Worksheet) -> list[dict[str, Any]]:
        labels = [_text(cell.value) for cell in sheet[1]]
        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(dict(zip(labels, values)))
        return rows

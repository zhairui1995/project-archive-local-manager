from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from database import Database
from services.excel_service import ExcelService
from services.file_service import FileService


def test_excel_import_and_export(tmp_path: Path) -> None:
    database = Database(tmp_path / "excel.db")
    database.initialize()
    files = FileService(database)
    excel = ExcelService(database, files)
    project_id = files.create_project({"name": "Excel 项目"})

    input_path = tmp_path / "input.xlsx"
    workbook = load_workbook(
        Path(__file__).resolve().parents[1] / "deliverables" / "档案迁移导入模板.xlsx"
    )
    sheet = workbook["档案导入"]
    sheet.append(["X-01", "导入合同", "纸质", "2份", 2, 0, ""])
    workbook.save(input_path)
    result = excel.import_project(project_id, input_path)
    assert result == {"imported": 1, "failed": 0, "errors": []}

    output_path = excel.export_project(project_id, tmp_path / "output.xlsx")
    exported_book = load_workbook(output_path, data_only=True)
    exported = exported_book["档案台账"]
    headers = {cell.value: index for index, cell in enumerate(exported[1], start=1)}
    assert exported.cell(2, headers["项目名称"]).value == "Excel 项目"
    assert exported.cell(2, headers["当前状态"]).value == "在库"


def test_export_and_import_all_round_trip(tmp_path: Path) -> None:
    source_db = Database(tmp_path / "source.db")
    source_db.initialize()
    source_files = FileService(source_db)
    source_excel = ExcelService(source_db, source_files)
    parent = source_files.create_project({"name": "大项目", "construction_company": "施工单位甲"})
    child = source_files.create_project(
        {"name": "消防子项目", "short_name": "消防", "parent_project_id": parent}
    )
    file_id = source_files.create_file(
        {
            "project_id": child,
            "name": "验收资料",
            "original_count": 2,
            "copy_count": 1,
        }
    )
    from services.borrow_service import BorrowService

    BorrowService(source_db).borrow_file(
        file_id, "张三", media_type="原件", quantity=1, borrow_date="2026-06-20"
    )
    path = source_excel.export_all(tmp_path / "all.xlsx")

    target_db = Database(tmp_path / "target.db")
    target_db.initialize()
    target_files = FileService(target_db)
    result = ExcelService(target_db, target_files).import_all(path)
    assert result == {"projects": 2, "files": 1, "borrows": 1}
    restored = target_files.get_file(file_id)
    assert restored["available_original"] == 1
    assert target_files.get_project(child)["parent_project_id"] == parent


def test_delivered_migration_template_is_importable(tmp_path: Path) -> None:
    database = Database(tmp_path / "template.db")
    database.initialize()
    files = FileService(database)
    excel = ExcelService(database, files)
    project_id = files.create_project({"name": "模板导入测试"})

    source = Path(__file__).resolve().parents[1] / "deliverables" / "档案迁移导入模板.xlsx"
    target = tmp_path / "filled-template.xlsx"
    workbook = load_workbook(source)
    sheet = workbook["档案导入"]
    sheet.append(["A-101", "模板合同", "纸质+电子", "2份", 2, 1, ""])
    workbook.save(target)

    result = excel.import_project(project_id, target)
    assert result == {"imported": 1, "failed": 0, "errors": []}
    imported = files.list_files(project_id)[0]
    assert imported["file_name"] == "模板合同"
    assert imported["original_count"] == 2
    assert imported["copy_count"] == 1
